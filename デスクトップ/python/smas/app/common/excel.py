import openpyxl as xl
from openpyxl.drawing.image import Image
import pandas as pd

import config
import common


##エクセル操作##


#ファイルのシート名を全て取得
def get_sheet_names(path):
     #ブックを取得
    wb = xl.load_workbook(path)
     #シートを取得 
    sheets = wb.sheetnames
    return sheets

#シート名変更
def change_sheet_name(path,name):
	#ブックを取得
    wb = xl.load_workbook(path)
    ws = wb.worksheets[0]
    #シート名の変更
    ws.title = name


#画像を貼り付ける
def paste_image(excel_path,img_path,position):
	flg = True
	try:
		#ブックを取得
		wb = xl.load_workbook(excel_path)
		ws = wb.worksheets[0]
		#挿入する画像を指定
		img = Image(img_path)
		#画像挿入
		ws.add_image(img, position)
		#保存
		wb.save(excel_path)
		wb.close()
	except FileNotFoundError as e:
		print("ファイルが見つかりません", e)
		flg = False
	except Exception as e:
		print("ファイルパスを確認してください", e)
		flg = False
	return flg		

#グラフを貼り付けたエクセルファイル出力
def export_excel(fig,png_path,excel_path,col_del,sheet_name,index_flg,df):
	#PNG作成
	common.graph.generate_png(fig,png_path)
	#不要なカラムを削除
	if col_del:
		df.drop(columns=col_del,inplace=True)
	#エクセル出力
	df.to_excel(excel_path, sheet_name=sheet_name,index=index_flg)
	position = "B" + str(len(df) + 5)
	flg = paste_image(excel_path,png_path,position)
	return flg
